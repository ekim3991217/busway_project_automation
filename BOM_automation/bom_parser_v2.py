#!/usr/bin/env python3
"""
bom_parser.py

Usage:
  python bom_parser.py

What it does:
  1) Prompts user to paste an Excel file path.
  2) Converts ALL cells in ALL sheets to values-only (removes formulas/dependencies).
  3) Deletes worksheet named "REFERENCE" (case-insensitive match).
  4) On the "BOM" sheet, UNMERGES all merged cells and resets horizontally centered cells to 'general'.
  5) On the "BOM" sheet, removes columns that are completely empty between rows 7–77 (inclusive), limited to J:AT.
  6) On the "BOM" sheet, removes rows 1:3.
  7) Writes an edited copy next to the original with "_editedBOM" appended to the filename.
"""

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.cell.cell import MergedCell

def is_cell_empty(value):
    """Treat None or whitespace-only strings as empty."""
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def main():
    # --- Filename prompt section ---
    file_input = input("COPY AND PASTE EXCEL FILEPATH HERE: ").strip('"').strip()
    src_path = Path(file_input).expanduser().resolve()

    if not src_path.exists():
        print(f"[ERROR] File not found: {src_path}")
        sys.exit(1)
    if src_path.suffix.lower() != ".xlsx":
        print(f"[ERROR] This script expects a .xlsx file. Got: {src_path.suffix}")
        sys.exit(1)

    # --- Workbook processing ---
    try:
        # Load with formulas/styles
        wb = load_workbook(filename=str(src_path), data_only=False)
        # Load with cached computed values
        wb_vals = load_workbook(filename=str(src_path), data_only=True)
    except Exception as e:
        print(f"[ERROR] Failed to open workbook: {e}")
        sys.exit(1)

    # ========================= Freeze formulas to values (safe for merged cells) =========================
    replaced = 0
    for name in wb.sheetnames:
        if name not in wb_vals.sheetnames:
            continue

        ws_main = wb[name]      # workbook with formulas/styles
        ws_vals = wb_vals[name] # workbook exposing cached values

        # Use max dimensions to be safe
        mr = max(ws_main.max_row, ws_vals.max_row)
        mc = max(ws_main.max_column, ws_vals.max_column)

        for r in range(1, mr + 1):
            for c in range(1, mc + 1):
                cell_main = ws_main.cell(row=r, column=c)

                # Non-anchor cells in merged regions are read-only -> skip them
                if isinstance(cell_main, MergedCell):
                    continue

                val = ws_vals.cell(row=r, column=c).value

                if val is not None:
                    # Overwrite with cached value (this removes formula/dependency)
                    cell_main.value = val
                    replaced += 1
                else:
                    # Fallback: if Excel didn't cache a value, leave the original content in place.
                    # (If you prefer to blank formulas without cached values, you could set to None.)
                    pass

    print(f"[INFO] Converted cells to values-only (updated {replaced} cell(s)).")
    # =====================================================================================================

    # 1) Delete worksheet "REFERENCE" (case-insensitive)
    ref_sheet_name = None
    for name in wb.sheetnames:
        if name.casefold() == "reference":
            ref_sheet_name = name
            break
    if ref_sheet_name:
        del wb[ref_sheet_name]
        print(f"[INFO] Deleted worksheet: {ref_sheet_name}")
    else:
        print("[INFO] No 'REFERENCE' worksheet found (skipping).")

    # 2) & 3) Operate on "BOM" sheet
    bom_sheet_name = None
    for name in wb.sheetnames:
        if name.casefold() == "bom":
            bom_sheet_name = name
            break

    if not bom_sheet_name:
        print("[ERROR] 'BOM' worksheet not found. No changes made to data sheets.")
        sys.exit(1)

    ws = wb[bom_sheet_name]

    # Unmerge all merged ranges
    merged_count = len(ws.merged_cells.ranges)
    if merged_count:
        for m in list(ws.merged_cells.ranges):
            ws.unmerge_cells(range_string=m.coord)
        print(f"[INFO] Unmerged {merged_count} merged cell range(s).")
    else:
        print("[INFO] No merged cells found to unmerge.")

    # Reset horizontally centered cells to 'general' (keep vertical alignment as-is)
    center_fixed = 0
    max_row = ws.max_row
    max_col = ws.max_column
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            al = cell.alignment
            # If horizontal is 'center' or 'centerContinuous', reset to general
            if al and al.horizontal and ("center" in al.horizontal):
                cell.alignment = Alignment(horizontal="general", vertical=al.vertical)
                center_fixed += 1
    print(f"[INFO] Reset horizontal centering on {center_fixed} cell(s).")

    # Remove columns empty ONLY within rows 7..77 (inclusive), and ONLY for J:AT
    window_start = 7
    window_end = min(77, ws.max_row)  # don't exceed actual last row
    cols_deleted = []

    if ws.max_row >= window_start and ws.max_column >= 10:
        # Iterate right->left; J=10, AT=46
        for col_idx in range(min(ws.max_column, 46), 9, -1):
            empty_in_window = True
            for row_idx in range(window_start, window_end + 1):
                if not is_cell_empty(ws.cell(row=row_idx, column=col_idx).value):
                    empty_in_window = False
                    break
            if empty_in_window:
                ws.delete_cols(col_idx, 1)
                cols_deleted.append(get_column_letter(col_idx))
    else:
        print("[INFO] No columns beyond I (J+) to evaluate or not enough rows to check 7–77 window.")

    if cols_deleted:
        print(f"[INFO] Deleted columns empty in rows 7–77: {', '.join(reversed(cols_deleted))}")
    else:
        print("[INFO] No columns empty across rows 7–77 found to delete.")

    # Remove rows 1:3
    ws.delete_rows(idx=1, amount=3)
    print("[INFO] Deleted rows 1:3 on 'BOM'.")

    # Save edited file alongside original
    out_path = src_path.with_name(f"{src_path.stem}_editedBOM{src_path.suffix}")
    try:
        wb.save(str(out_path))
    except Exception as e:
        print(f"[ERROR] Failed to save edited workbook: {e}")
        sys.exit(1)

    print(f"[SUCCESS] Edited workbook saved to:\n{out_path}")


if __name__ == "__main__":
    main()
