#!/usr/bin/env python3
"""
ps_cleaner.py

Usage (preferred):
  python ps_cleaner.py "C:\path\to\250826 PS-USA-2444 EGLINTON PJT-A.xlsx"

Also works interactively if no arg is provided:
  python ps_cleaner.py
  (then paste path when prompted)

What it does:
  1) Deletes worksheet "COVER" (case-insensitive).
  2) Converts all cells to values-only (removes formulas/dependencies).
  3) On "PS" sheet, clears columns L and beyond (content + cell styles), without deleting columns.
  4) Saves a copy with file name prefix changed from "yymmdd PS-" to "yymmdd RFP-",
     preserving the original extension and formatting as much as openpyxl allows.
     For .xlsm sources, macros are preserved.
"""

import sys
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border
from openpyxl.cell.cell import MergedCell


def get_src_path():
    """Get the source file path from argv or an interactive prompt."""
    if len(sys.argv) > 1:
        raw = sys.argv[1]
    else:
        raw = input("COPY AND PASTE EXCEL FILEPATH HERE: ")
    p = Path(raw.strip().strip('"').strip("'")).expanduser().resolve()
    if not p.exists():
        print(f"[ERROR] File not found: {p}")
        sys.exit(1)
    if p.suffix.lower() not in (".xlsx", ".xlsm"):
        print(f"[ERROR] Expected a .xlsx or .xlsm file. Got: {p.suffix}")
        sys.exit(1)
    return p


def rename_to_rfp(src_path: Path) -> Path:
    """
    Replace leading 'yymmdd PS-' with 'yymmdd RFP-'.
    If the exact pattern isn't found, fall back to inserting 'RFP-' after date if present,
    else append '_RFP' before extension.
    """
    name = src_path.name  # includes extension
    # Primary pattern: 6 digits, space, PS-
    m = re.match(r"^(\d{6})\sPS-(.*)$", name, flags=re.IGNORECASE)
    if m:
        new_name = f"{m.group(1)} RFP-{m.group(2)}"
        return src_path.with_name(new_name)

    # Secondary: if it starts with date but not exact " PS-"
    m2 = re.match(r"^(\d{6})(\s?)(.*)$", name)
    if m2 and "RFP-" not in name.upper():
        new_name = f"{m2.group(1)} RFP-{m2.group(3)}"
        return src_path.with_name(new_name)

    # Fallback: append _RFP before extension
    return src_path.with_name(f"{src_path.stem}_RFP{src_path.suffix}")


def freeze_to_values(wb, wb_vals):
    """
    Overwrite each writable (non-merged) cell's value with the cached data_only value.
    Leaves styles/formatting intact. Skips non-anchor merged cells to avoid write errors.
    """
    replaced = 0
    for name in wb.sheetnames:
        if name not in wb_vals.sheetnames:
            continue
        ws_main = wb[name]
        ws_vals = wb_vals[name]
        mr = max(ws_main.max_row, ws_vals.max_row)
        mc = max(ws_main.max_column, ws_vals.max_column)
        for r in range(1, mr + 1):
            for c in range(1, mc + 1):
                cell_main = ws_main.cell(row=r, column=c)
                if isinstance(cell_main, MergedCell):
                    continue
                val = ws_vals.cell(row=r, column=c).value
                if val is not None:
                    cell_main.value = val
                    replaced += 1
                else:
                    # If Excel didn't cache a value, keep original (formula/text) as-is.
                    pass
    print(f"[INFO] Converted cells to values-only (updated {replaced} cell(s)).")


def clear_PS_from_L_right(ws_ps):
    """
    On the PS sheet, clear all content and styles from column L onward,
    without deleting columns (so layout/widths aren’t shifted).
    Also unmerges any merged ranges that intersect columns >= L to avoid write errors.
    """
    if ws_ps is None:
        print("[ERROR] 'PS' worksheet not found.")
        sys.exit(1)

    L_COL_IDX = 12  # Column L
    if ws_ps.max_column < L_COL_IDX:
        print("[INFO] 'PS' has no columns at or beyond L. Skipping clear.")
        return

    # Unmerge any merged ranges touching columns >= L
    merged_ranges = list(ws_ps.merged_cells.ranges)
    unmerged = 0
    for rng in merged_ranges:
        if rng.max_col >= L_COL_IDX:  # any overlap with L+
            ws_ps.unmerge_cells(range_string=rng.coord)
            unmerged += 1
    if unmerged:
        print(f"[INFO] Unmerged {unmerged} merged range(s) intersecting columns L+ on 'PS'.")

    mr = ws_ps.max_row
    mc = ws_ps.max_column

    # Reset styles to default + clear values/comments/hyperlinks for L..end
    cleared = 0
    for c in range(L_COL_IDX, mc + 1):
        for r in range(1, mr + 1):
            cell = ws_ps.cell(row=r, column=c)
            # Clear content
            cell.value = None
            cell.hyperlink = None
            cell.comment = None
            # Reset style to defaults
            cell.font = Font()                 # default font
            cell.fill = PatternFill()          # no fill
            cell.border = Border()             # no border
            cell.alignment = Alignment(horizontal="general")  # default alignment
            cell.number_format = "General"     # default number format
            cleared += 1

    print(f"[INFO] Cleared values/styles from column L to {get_column_letter(mc)} on 'PS' "
          f"({cleared} cell(s)).")


def main():
    src_path = get_src_path()
    keep_vba = src_path.suffix.lower() == ".xlsm"

    # Load workbooks. For .xlsm, keep_vba=True to preserve macros.
    try:
        wb = load_workbook(filename=str(src_path), data_only=False, keep_vba=keep_vba)
        wb_vals = load_workbook(filename=str(src_path), data_only=True)
    except Exception as e:
        print(f"[ERROR] Failed to open workbook: {e}")
        sys.exit(1)

    # 1) Delete COVER (case-insensitive)
    cover_name = next((n for n in wb.sheetnames if n.casefold() == "cover"), None)
    if cover_name:
        del wb[cover_name]
        print(f"[INFO] Deleted worksheet: {cover_name}")
    else:
        print("[INFO] No 'COVER' worksheet found (skipping).")

    # 2) Convert to values-only on remaining sheets
    freeze_to_values(wb, wb_vals)

    # 3) Clear columns L+ on PS
    ps_name = next((n for n in wb.sheetnames if n.casefold() == "ps"), None)
    ws_ps = wb[ps_name] if ps_name else None
    clear_PS_from_L_right(ws_ps)

    # 4) Build output filename (yymmdd PS- -> yymmdd RFP-), keep extension (.xlsx/.xlsm)
    out_path = rename_to_rfp(src_path)
    try:
        wb.save(str(out_path))
    except Exception as e:
        print(f"[ERROR] Failed to save workbook: {e}")
        sys.exit(1)

    print(f"[SUCCESS] Saved edited file to:\n{out_path}")


if __name__ == "__main__":
    main()