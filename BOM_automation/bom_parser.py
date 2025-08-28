#!/usr/bin/env python3
"""
bom_parser.py

Usage:
  python bom_parser.py "C:\path\to\your\file.xlsx"

What it does:
  1) Deletes worksheet named "REFERENCE" (case-insensitive match).
  2) On the "BOM" sheet, removes columns that are completely empty from row 7 to the last used row.
  3) On the "BOM" sheet, removes rows 1:3.
  4) Writes an edited copy next to the original with "_editedBOM" appended to the filename.
"""

import sys
import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def is_cell_empty(value):
    """Treat None or whitespace-only strings as empty."""
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def main():
    parser = argparse.ArgumentParser(description="Edit a BOM Excel file.")
    parser.add_argument("filepath", help="Path to the .xlsx file to edit")
    args = parser.parse_args()

    src_path = Path(args.filepath).expanduser().resolve()
    if not src_path.exists():
        print(f"[ERROR] File not found: {src_path}")
        sys.exit(1)
    if src_path.suffix.lower() != ".xlsx":
        print(f"[ERROR] This script expects a .xlsx file. Got: {src_path.suffix}")
        sys.exit(1)

    # Load workbook with formatting preserved as much as openpyxl allows
    try:
        wb = load_workbook(filename=str(src_path), data_only=False)
    except Exception as e:
        print(f"[ERROR] Failed to open workbook: {e}")
        sys.exit(1)

    # 1) Delete worksheet "REFERENCE" (case-insensitive)
    ref_sheet_name = None
    for name in wb.sheetnames:
        if name.casefold() == "reference":
            ref_sheet_name = name
            break
    if ref_sheet_name:
        del wb[ref_sheet_name]
        print(f"[INFO] Deleted worksheet: {ref_sheet_name}")
    else:
        print("[INFO] No 'REFERENCE' worksheet found (skipping).")

    # 2) & 3) Operate on "BOM" sheet
    bom_sheet_name = None
    for name in wb.sheetnames:
        if name.casefold() == "bom":
            bom_sheet_name = name
            break

    if not bom_sheet_name:
        print("[ERROR] 'BOM' worksheet not found. No changes made to data sheets.")
        sys.exit(1)

    ws = wb[bom_sheet_name]

    # Remove columns that are completely empty from row 7 down
    # Iterate from right to left to avoid index shifting
    max_row = ws.max_row
    max_col = ws.max_column

    cols_deleted = []
    for col_idx in range(max_col, 0, -1):
        empty_from_row7 = True
        for row_idx in range(7, max_row + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if not is_cell_empty(cell_val):
                empty_from_row7 = False
                break
        if empty_from_row7:
            ws.delete_cols(col_idx, 1)
            cols_deleted.append(get_column_letter(col_idx))

    if cols_deleted:
        print(f"[INFO] Deleted empty columns (from row 7 down): {', '.join(reversed(cols_deleted))}")
    else:
        print("[INFO] No fully empty columns (row 7→end) found to delete.")

    # Remove rows 1:3
    ws.delete_rows(idx=1, amount=3)
    print("[INFO] Deleted rows 1:3 on 'BOM'.")

    # 4) Save edited file alongside original
    out_path = src_path.with_name(f"{src_path.stem}_editedBOM{src_path.suffix}")
    try:
        wb.save(str(out_path))
    except Exception as e:
        print(f"[ERROR] Failed to save edited workbook: {e}")
        sys.exit(1)

    print(f"[SUCCESS] Edited workbook saved to:\n{out_path}")


if __name__ == "__main__":
    main()
